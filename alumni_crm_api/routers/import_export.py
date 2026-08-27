import csv
import datetime
import io
import logging

import openpyxl
import pg8000.dbapi
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse

from database import get_db
from routers.etudiants import (
    ACADEMIC_EMAIL_DOMAIN,
    _email_academique_exists,
    _validate_availability_status,
)
from security import require_admin_api_key
from utils import normalize_academic_slug
from config import settings

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/import",
    tags=["Import / Export"],
    dependencies=[Depends(require_admin_api_key)],
)

COLUMN_MAP = {
    "prenom": "prenom",
    "nom": "nom",
    "email": "email",
    "telephone": "telephone",
    "promotion": "promotion",
    "annee_diplome": "annee_diplome",
    "entreprise": "entreprise",
    "poste": "poste",
    "secteur": "secteur",
    "entreprise_pays": "entreprise_pays",
    "entreprise_ville": "entreprise_ville",
    "linkedin": "linkedin",
    "adresse": "adresse",
    "ville": "ville",
    "pays": "pays",
    "statut_disponibilite": "statut_disponibilite",
    "competences": "competences",
    "date_naissance": "date_naissance",
    "date_inscription": "date_inscription",
    "email_academique": "email_academique",
    "parcours_anterieur": "parcours_anterieur",
    "type_contrat": "type_contrat",
    "date_debut": "date_debut",
    "date_fin": "date_fin",
    "poste_actuel": "poste_actuel",
}

_FALSY_POSTE_ACTUEL = {"false", "0", "non", "faux", "no", "f", "n"}

_DATE_FORMATS_JJ = ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y")


def _detecter_separateur_csv(text: str) -> str:
    """Détecte le séparateur d'un CSV importé (Excel FR utilise ';',
    les exports standard ','). On choisit le délimiteur qui produit le
    plus de colonnes sur la ligne d'en-tête ; fallback virgule."""
    lignes = text.splitlines()
    premiere_ligne = lignes[0] if lignes else ""
    meilleur, max_champs = ",", 1
    for delim in (",", ";", "\t"):
        try:
            champs = next(csv.reader([premiere_ligne], delimiter=delim))
        except (StopIteration, csv.Error):
            continue
        if len(champs) > max_champs:
            meilleur, max_champs = delim, len(champs)
    return meilleur


def _parse_date_cell(value):
    """Interprète une cellule de date d'un fichier importé.

    Accepte : cellule date Excel native (datetime/date), chaîne ISO
    (AAAA-MM-JJ, avec ou sans heure) et chaîne JJ/MM/AAAA. Renvoie None
    si la cellule est vide. Lève ValueError si non vide mais illisible,
    ce qui rejette la ligne sans faire échouer tout l'import.
    """
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        return datetime.date.fromisoformat(text[:10])
    except ValueError:
        pass
    for fmt in _DATE_FORMATS_JJ:
        try:
            return datetime.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(
        f"Date invalide : '{text}' (formats acceptés : AAAA-MM-JJ ou JJ/MM/AAAA)."
    )


def _parse_poste_actuel(value: str) -> bool:
    """Interprète la colonne poste_actuel d'un fichier importé.

    Valeurs reconnues comme fausses : false, 0, non, faux, no, f, n
    (insensible à la casse). Vide ou absente → True (comportement historique).
    """
    return value.strip().lower() not in _FALSY_POSTE_ACTUEL


def _resolve_promotion(cursor, promotion_name: str | None, annee_diplome: int | None):
    if promotion_name:
        cursor.execute(
            "SELECT id_promotion FROM PROMOTION WHERE nom_promotion = %s",
            (promotion_name,),
        )
        row = cursor.fetchone()
        if row:
            return row[0]
    if annee_diplome:
        cursor.execute(
            "SELECT id_promotion FROM PROMOTION WHERE annee_diplome = %s ORDER BY id_promotion LIMIT 1",
            (annee_diplome,),
        )
        row = cursor.fetchone()
        if row:
            return row[0]
    cursor.execute("SELECT id_promotion FROM PROMOTION ORDER BY id_promotion LIMIT 1")
    row = cursor.fetchone()
    return row[0] if row else None


def _resolve_entreprise(cursor, nom_entreprise: str | None, secteur: str | None,
                        pays: str = "", ville: str = ""):
    if not nom_entreprise:
        return None
    cursor.execute(
        "SELECT id_entreprise FROM ENTREPRISE WHERE nom_entreprise = %s",
        (nom_entreprise,),
    )
    row = cursor.fetchone()
    if row:
        # Entreprise existante : on ne touche jamais à ses coordonnées,
        # pour ne pas dégrader un pays/ville déjà renseigné avec des
        # valeurs vides ou différentes du fichier importé.
        return row[0]
    cursor.execute(
        """INSERT INTO ENTREPRISE (nom_entreprise, secteur_activite, pays, ville)
           VALUES (%s, %s, %s, %s) RETURNING id_entreprise""",
        (nom_entreprise, secteur or "Autre", pays or "", ville or ""),
    )
    return cursor.fetchone()[0]


def _resolve_email_academique_import(
    cursor,
    prenom: str,
    nom: str,
    provided: str | None,
    used_in_file: set,
) -> str | None:
    """Résout email_academique pour une ligne importée.

    Reprend la logique de l'inscription individuelle (routers/etudiants.py)
    en l'étendant aux doublons internes d'un même fichier : la recherche
    d'une valeur libre tient compte à la fois de la base ETUDIANT et des
    emails déjà acceptés dans les lignes précédentes (traitement dans
    l'ordre du fichier, d'où une numérotation déterministe).
    """
    slug_prenom = normalize_academic_slug(prenom)
    slug_nom = normalize_academic_slug(nom)
    base = f"{slug_prenom}.{slug_nom}" if slug_prenom and slug_nom else ""
    candidate = (provided or "").strip().lower()

    def _taken(email: str) -> bool:
        return email in used_in_file or _email_academique_exists(cursor, email)

    if base:
        auto_generated = f"{base}@{ACADEMIC_EMAIL_DOMAIN}"
        if not candidate or candidate == auto_generated:
            candidate = auto_generated
            suffix = 2
            while _taken(candidate):
                candidate = f"{base}{suffix}@{ACADEMIC_EMAIL_DOMAIN}"
                suffix += 1
            return candidate

    if candidate and _taken(candidate):
        raise ValueError(
            f"L'email académique '{candidate}' est déjà utilisé par un autre "
            "étudiant ou par une autre ligne du fichier."
        )
    return candidate or None


@router.post("/excel")
async def import_excel(file: UploadFile = File(...), db=Depends(get_db)):
    if not file.filename or not file.filename.endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(status_code=400, detail="Format de fichier non supporté. Utilisez .xlsx ou .csv.")

    max_bytes = settings.max_upload_size_mb * 1024 * 1024
    content = await file.read()
    if len(content) > max_bytes:
        raise HTTPException(
            status_code=413,
            detail=f"Fichier trop volumineux. Taille maximale : {settings.max_upload_size_mb} Mo.",
        )
    try:
        if file.filename.lower().endswith(".csv"):
            text = content.decode("utf-8-sig")
            rows = list(csv.reader(io.StringIO(text), delimiter=_detecter_separateur_csv(text)))
        else:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
    except Exception:
        raise HTTPException(status_code=400, detail="Impossible de lire le fichier.")

    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Le fichier est vide ou ne contient pas de données.")

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    col_map = {}
    for idx, header in enumerate(headers):
        normalized = header.replace(" ", "_")
        if normalized in COLUMN_MAP:
            col_map[normalized] = idx

    if "nom" not in col_map and "prenom" not in col_map:
        raise HTTPException(status_code=400, detail="Colonnes 'nom' et/ou 'prenom' introuvables dans le fichier.")

    imported = 0
    errors = []
    used_in_file = set()
    cursor = db.cursor()

    try:
        for row_num, row in enumerate(rows[1:], start=2):
            cursor.execute("SAVEPOINT sp_row;")
            try:
                def get_val(col_name):
                    idx = col_map.get(col_name)
                    return str(row[idx]).strip() if idx is not None and idx < len(row) and row[idx] else ""

                def get_raw(col_name):
                    # Valeur brute de la cellule (conserve les types natifs
                    # openpyxl, ex : datetime pour une cellule au format date).
                    idx = col_map.get(col_name)
                    return row[idx] if idx is not None and idx < len(row) else None

                nom = get_val("nom")
                prenom = get_val("prenom")
                email = get_val("email")

                if not nom or not prenom:
                    errors.append({"row": row_num, "message": "Nom ou prénom manquant."})
                    continue

                promotion_name = get_val("promotion") or None
                annee_diplome_str = get_val("annee_diplome")
                annee_diplome = int(annee_diplome_str) if annee_diplome_str.isdigit() else None
                id_promotion = _resolve_promotion(cursor, promotion_name, annee_diplome)

                if not id_promotion:
                    errors.append({"row": row_num, "message": "Aucune promotion trouvée."})
                    continue

                telephone = get_val("telephone") or ""
                email_val = email or f"{prenom.lower()}.{nom.lower()}@placeholder.com"

                date_naissance_val = get_val("date_naissance") or "2000-01-01"
                # Dates réelles du fichier, fallback sur la date du jour si
                # la cellule est vide (comportement historique préservé).
                date_inscription_val = _parse_date_cell(get_raw("date_inscription"))
                if date_inscription_val is None:
                    date_inscription_val = datetime.date.today()
                date_debut_val = _parse_date_cell(get_raw("date_debut"))
                if date_debut_val is None:
                    date_debut_val = datetime.date.today()
                type_contrat_val = get_val("type_contrat") or "Non renseigné"
                poste_actuel_val = _parse_poste_actuel(get_val("poste_actuel"))
                # Un poste actuel n'a pas de date de fin : toute valeur fournie
                # est ignorée. Sinon la date est acceptée (None si absente).
                date_fin_val = (
                    None if poste_actuel_val else _parse_date_cell(get_raw("date_fin"))
                )
                if date_fin_val is not None and date_fin_val < date_debut_val:
                    raise ValueError(
                        f"date_fin ({date_fin_val}) antérieure à "
                        f"date_debut ({date_debut_val})."
                    )
                email_academique_val = get_val("email_academique") or None
                email_academique_val = _resolve_email_academique_import(
                    cursor, prenom, nom, email_academique_val, used_in_file
                )
                parcours_val = get_val("parcours_anterieur") or ""
                address_val = get_val("adresse") or ""
                city_val = get_val("ville") or ""
                country_val = get_val("pays") or ""
                linkedin_val = get_val("linkedin") or ""
                availability_val = get_val("statut_disponibilite") or ""
                # Même validation que PUT/PATCH : une valeur fournie mais
                # invalide rejette la ligne (sans faire échouer tout l'import).
                if availability_val:
                    try:
                        _validate_availability_status(availability_val)
                    except HTTPException as exc:
                        errors.append({"row": row_num, "message": str(exc.detail)})
                        continue
                competences_val = get_val("competences") or ""
                skills_list = [s.strip() for s in competences_val.split(",") if s.strip()] if competences_val else []

                cursor.execute(
                    """INSERT INTO ETUDIANT (nom, prenom, email, email_academique, telephone,
                       date_naissance, parcours_anterieur, date_inscription, id_promotion,
                       address, city, country, linkedin, availability_status, skills)
                       VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s,
                       %s, %s, %s, %s, %s, %s::jsonb) RETURNING id_etudiant""",
                    (nom, prenom, email_val, email_academique_val, telephone,
                     date_naissance_val, parcours_val, date_inscription_val, id_promotion,
                     address_val, city_val, country_val, linkedin_val, availability_val,
                     str(__import__('json').dumps(skills_list))),
                )
                id_etudiant = cursor.fetchone()[0]

                entreprise_nom = get_val("entreprise")
                if entreprise_nom:
                    secteur = get_val("secteur")
                    id_entreprise = _resolve_entreprise(
                        cursor, entreprise_nom, secteur,
                        get_val("entreprise_pays"), get_val("entreprise_ville"),
                    )
                    poste = get_val("poste") or "Non renseigné"
                    cursor.execute(
                        """INSERT INTO EXPERIENCE_PRO
                           (intitule_poste, type_contrat, date_debut, date_fin, salaire,
                            poste_actuel, id_entreprise, id_etudiant)
                           VALUES (%s, %s, %s, %s, 0, %s, %s, %s)""",
                        (poste, type_contrat_val, date_debut_val, date_fin_val,
                         poste_actuel_val, id_entreprise, id_etudiant),
                    )

                imported += 1
                if email_academique_val:
                    used_in_file.add(email_academique_val)
            except ValueError as exc:
                errors.append({"row": row_num, "message": str(exc)})
                continue
            except Exception:
                errors.append({"row": row_num, "message": "Erreur inattendue lors du traitement de la ligne."})
                try:
                    cursor.execute("ROLLBACK TO SAVEPOINT sp_row;")
                except Exception:
                    pass
                continue

        db.commit()
    finally:
        cursor.close()

    return {
        "message": f"Import terminé : {imported} alumni importé(s), {len(errors)} erreur(s).",
        "imported": imported,
        "errors": errors,
    }


@router.get("/template")
async def download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Modèle import"
    # Les en-têtes sont dérivés de COLUMN_MAP pour garantir qu'un fichier
    # généré ici est toujours reconnu tel quel par /import/excel.
    headers = list(COLUMN_MAP.keys())
    ws.append(headers)
    ws.append([
        "Jean", "Dupont", "jean.dupont@email.com", "0612345678", "Promo 2024", 2024,
        "Acme Corp", "Développeur", "Technologie", "France", "Paris",
        "https://linkedin.com/in/jeandupont",
        "15 Rue de Paris", "Paris", "France", "en_poste", "Python, React",
        "1995-06-15", "2020-09-01", "jean.dupont@univ.fr", "Licence Informatique",
        "CDI", "2024-10-01", "", "Oui",
    ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=modele_import_alumni.xlsx"},
    )


@router.get("/export/alumni")
async def export_alumni(db=Depends(get_db)):
    cursor = db.cursor()
    try:
        cursor.execute("""
            SELECT e.prenom, e.nom, e.email, e.telephone, p.nom_promotion,
                   p.annee_diplome, ent.nom_entreprise, exp.intitule_poste,
                   ent.secteur_activite, ent.pays, ent.ville,
                   e.linkedin, e.address, e.city, e.country,
                   e.availability_status, e.skills, e.date_naissance,
                   e.date_inscription, e.email_academique, e.parcours_anterieur,
                   exp.type_contrat, exp.date_debut, exp.date_fin, exp.poste_actuel
            FROM ETUDIANT e
            LEFT JOIN PROMOTION p ON e.id_promotion = p.id_promotion
            LEFT JOIN EXPERIENCE_PRO exp ON e.id_etudiant = exp.id_etudiant AND exp.poste_actuel = TRUE
            LEFT JOIN ENTREPRISE ent ON exp.id_entreprise = ent.id_entreprise
            ORDER BY e.nom, e.prenom
        """)
        rows = cursor.fetchall()
    finally:
        cursor.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alumni"
    # En-têtes alignés sur COLUMN_MAP : le fichier produit par cet export peut
    # être réimporté tel quel via /import/excel sans perte ni renommage.
    headers = [
        "prenom", "nom", "email", "telephone", "promotion", "annee_diplome",
        "entreprise", "poste", "secteur", "entreprise_pays", "entreprise_ville",
        "linkedin", "adresse", "ville", "pays",
        "statut_disponibilite", "competences", "date_naissance",
        "date_inscription", "email_academique", "parcours_anterieur",
        "type_contrat", "date_debut", "date_fin", "poste_actuel",
    ]
    ws.append(headers)
    import json as _json
    for row in rows:
        flat = []
        for cell in row:
            if isinstance(cell, list):
                flat.append(", ".join(str(x) for x in cell))
            elif cell is None:
                flat.append("")
            else:
                flat.append(str(cell))
        ws.append(flat)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=alumni_export.xlsx"},
    )
